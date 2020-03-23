import ssl
import time
import OpenSSL
from lxml import objectify
import openpyxl
import os
import datetime
from openpyxl.styles import PatternFill
import requests


"""
# Проверка редиректа
def check_redirect(check_url):
    pattern_meta_redirect = "url='{0,2}?(https?://[-A-Za-z0-9+&@#/%?=~_|!:,.;]+[-A-Za-z0-9+&@#/%=~_|])'{0,2}?"
    pattern_meta_url = "url="
    pattern_window_location_href = "window.location.href\s?=\s?\"(https?://[-A-Za-z0-9+&@#/%?=~_|!:,.;]+[-A-Za-z0-9+&@#/%=~_|])\""
    # pattern_window_location_href = "window.location.href"
    redirect = ''
    try:
        r = requests.get(check_url, headers={"User-Agent": user_agent}, verify=False, timeout=10, stream=True)
        soup = BeautifulSoup(r.content, "html.parser")
        html_source = soup.find_all('meta', attrs={'http-equiv': 'refresh'})
        find_url_tag = re.findall(pattern_meta_url, str(html_source), flags=re.IGNORECASE)
        find_wlh_url = re.findall(pattern_window_location_href, str(soup.find_all('script')), flags=re.IGNORECASE)
        # print(r.status_code)
        # print(soup)
        # print(r)
        redirect = r.url
        if redirect != check_url:
            # print(redirect)
            print("Redirect {0} {1}".format(check_url, redirect))
            return redirect
        elif html_source:
            # find_url = re.findall(pattern_meta_redirect, str(html_source),flags=re.IGNORECASE)
            # print(find_wlh_url)
            # обнаружение нестандартного тега url, например, в виде javascript или base64
            if find_url_tag:
                find_url = re.findall(pattern_meta_redirect, str(html_source), flags=re.IGNORECASE)
                if find_url:
                    # cтандартный тег url
                    check_url = "".join(find_url)
                    # print("".join(find_url))
                    return check_url
                else:
                    print("check_redirect_1. Подозрительный редирект в метатеге", html_source)
                    return check_url

        # Подозрительный редирект через window.location.href
        elif find_wlh_url:
            # print("".join(find_wlh_url))
            print("check_redirect. Подозрительный редирект в javascript", soup.script)
            return check_url

        else:
            # print("check_redirect_2 - ", check_url)
            return check_url
    except Exception as e:
        print("check_redirect error: ", e)
        # except requests.exceptions.ConnectionError as e:
        # redirect = check_url
        return False
"""


# Заплатка для ошибки с протоколом
ssl._create_default_https_context = ssl._create_unverified_context


# Получение сертификата с хоста по порту 443
def get_data(hostname):
    print("\n\naddress: ", addr)
    Comment = ""
    Header = ""
    Certificate = ""
    notBefore = ""
    notAfter = ""
    Days_left = ""
    Signed = ""
    try:
        if hostname[1] == "443":
            URL = "https://" + hostname[0]
        else:
            URL = "http://" + hostname[0]
        print(URL)
        headers = {'headers': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:51.0) Gecko/20100101 Firefox/51.0'}
        page = requests.get(URL, headers=headers, timeout=10, verify=False)
        status_code = page.status_code
        print("status_code: ", status_code)
        page = page.text

        if page:
            if page.find('<title') != -1:
                Header = page[page.find('<title>') + 7: page.find('</title>')]
            else:
                if page.find('<TITLE') != -1:
                    Header = page[page.find('<TITLE') + 7: page.find('</TITLE>')]
                else:
                    Header = page
        else:
            Header = "status_code: " + str(status_code)

        Availability = "Доступен"
        if hostname[1] != "443":
            return [hostname[0] + ":" + hostname[1], Availability, Header, Certificate, notBefore, notAfter, Days_left,
                    Signed, Comment]
        else:
            try:
                certificate = ssl.get_server_certificate((hostname[0], 443))
                decode_data = OpenSSL.crypto.load_certificate(OpenSSL.crypto.FILETYPE_PEM, certificate)

                print("issuer: ", decode_data.get_issuer())
                print("notAfter: ",
                      time.mktime(time.strptime(decode_data.get_notAfter().decode('ascii'), '%Y%m%d%H%M%SZ')))
                print("notBefore: ",
                      time.mktime(time.strptime(decode_data.get_notBefore().decode('ascii'), '%Y%m%d%H%M%SZ')))

                if str(decode_data.get_issuer()).endswith("/CN=TAG-CA'>"):
                    Signed = "Да"
                else:
                    Signed = "Нет"
                notBefore = format_data(decode_data.get_notBefore())
                notAfter = format_data(decode_data.get_notAfter())
                Days_left = days_left(decode_data.get_notAfter())
                Comment = str(decode_data.get_issuer())
                Certificate = "Есть"
            except Exception as Ex:
                Certificate = "Нет"
                Comment = str(Ex)
                return [hostname[0] + ":" + hostname[1], Availability, Header, Certificate, notBefore, notAfter,
                        Days_left, Signed, Comment]
    except Exception as Ex:
        Availability = "Не доступен"
        Comment = str(Ex)
        return [hostname[0] + ":" + hostname[1], Availability, Header, Certificate, notBefore, notAfter, Days_left,
                Signed, Comment]

    return [hostname[0] + ":" + hostname[1], Availability, Header, Certificate, notBefore, notAfter, Days_left,
            Signed, Comment]


# Вычисляем оставшиеся дни до конца сертификата
def days_left(time_):
    time_ = time_.decode('ascii')
    time_ = time.strptime(time_, '%Y%m%d%H%M%SZ')
    time_ = time.mktime(time_)
    if (time_ - time.time()) > 0:
        days = (time_ - time.time())//86400
    else:
        days = 0
    return int(days)


# Форматирование даты
def format_data(time_):
    time_ = time_.decode('ascii')
    time_ = time.strptime(time_, '%Y%m%d%H%M%SZ')
    # data = str(time_[2]) + "-" + str(time_[1]) + "-" + str(time_[0])
    data = datetime.datetime(time_[0], time_[1], time_[2], time_[3], time_[4], time_[5])
    return data


# Получение списка хостов из отчета masscan
def get_hostname(file):
    addresses = []
    with open(file) as fobj:
        xml = fobj.read()

    root = objectify.fromstring(xml)

    for host in root.getchildren():
        if host.tag == "host":
            addresses.append([host.address.attrib["addr"], host.ports.port.attrib["portid"]])

    return addresses


# Запись результатов в файл
def Excel(results):
    # объект
    wb = openpyxl.Workbook()

    for result in results:
        sheet_name = result[0]
        rows = result[1]

        # название страницы
        ws = wb.create_sheet(sheet_name, 0)
        ws.title = sheet_name

        # циклом записываем данные
        x = 0   # Номер ячейки
        for row in rows:
            x += 1
            ws.append(row)
            if x != 1 and row[6] != "":
                if row[6] == 0:
                    ws["G" + str(x)].fill = redFill
                elif row[6] <= 180:
                    ws["G" + str(x)].fill = orangeFill

        # Ширина ячеек
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 50
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 150

    # сохранение файла в текущую директорию
    wb.save("results.xlsx")


# Цвета для ячеек
redFill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")
orangeFill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")

# Получаем список отчетов в папке
directory = "./reports/"
files = os.listdir(directory)
files = filter(lambda x: x.endswith('.xml'), files)
results = []


# Прокручиваем отчеты один за другим
for file in files:
    addresses = get_hostname("./reports/" + file)
    result = [["Адрес", "Доступность", "Заголовок", "Сертификат", "Действует от", "Годен до", "Осталось дней", "Подписан", "Примечание"]]
    # Стучимся к хостам, проверяем сертификаты и доступность
    for addr in addresses:
        result.append(get_data(addr))

    results.append([file, result])

Excel(results)
